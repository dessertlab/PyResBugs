from __future__ import annotations

import csv
import hashlib
import json
import os
import re
import shutil
import tempfile
from collections import Counter
from collections.abc import Iterable, Iterator, Mapping
from dataclasses import dataclass
from pathlib import Path
from typing import Any, TextIO
from urllib.parse import urlparse

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook

SHEET_NAME = "PyResBugs"
HEADERS = (
    "Bug_Description",
    "Bug_Type",
    "CVE-ID",
    "Commit_URL",
    "Commit_sha",
    "Dataset_input",
    "Diff_patch",
    "Fault Free Code",
    "Faulty Code",
    "Fixed_Method",
    "Impact",
    "Implementation-Level Description",
    "Contextual-Level Description",
    "High-Level Description",
    "Project",
    "Python_Version",
    "Test_File_Path",
    "Url",
    "Fault_Acronym",
)
OPTIONAL_FIELDS = frozenset(("Bug_Type", "CVE-ID", "Impact", "Test_File_Path"))
REQUIRED_FIELDS = tuple(field for field in HEADERS if field not in OPTIONAL_FIELDS)
IDENTITY_FIELDS = (
    "Project",
    "Commit_sha",
    "Fixed_Method",
    "Fault_Acronym",
    "Faulty Code",
)
SHA_PATTERN = re.compile(r"[0-9a-fA-F]{7,40}")
ID_PATTERN = re.compile(r"(?:PB-)?([0-9A-Fa-f]{6,64})")


class DatasetError(RuntimeError):
    """Base class for actionable dataset errors."""


class RecordNotFoundError(DatasetError):
    pass


class AmbiguousSelectorError(DatasetError):
    pass


class InvalidRecordError(DatasetError):
    pass


@dataclass(frozen=True)
class ValidationIssue:
    level: str
    code: str
    message: str
    excel_row: int | None = None

    def __str__(self) -> str:
        location = f" row {self.excel_row}" if self.excel_row is not None else ""
        return f"{self.level.upper()} [{self.code}]{location}: {self.message}"


@dataclass(frozen=True)
class Record:
    values: Mapping[str, Any]
    excel_row: int

    @property
    def ordinal(self) -> int:
        return self.excel_row - 1

    @property
    def id(self) -> str:
        identity = [self.values.get(field) for field in IDENTITY_FIELDS]
        encoded = json.dumps(identity, ensure_ascii=False, separators=(",", ":")).encode("utf-8")
        return f"PB-{hashlib.sha256(encoded).hexdigest()[:12].upper()}"

    def as_dict(self, include_id: bool = True) -> dict[str, Any]:
        result = dict(self.values)
        if include_id:
            return {"Bug_ID": self.id, **result}
        return result


class Dataset:
    """An editable view of the PyResBugs Excel workbook."""

    def __init__(self, path: str | Path):
        self.path = Path(path)
        if not self.path.is_file():
            raise DatasetError(f"dataset file does not exist: {self.path}")
        self.workbook: Workbook = load_workbook(self.path, read_only=False, data_only=False)
        if SHEET_NAME not in self.workbook.sheetnames:
            self.close()
            raise DatasetError(f"workbook has no {SHEET_NAME!r} sheet")
        self.sheet = self.workbook[SHEET_NAME]

    def close(self) -> None:
        self.workbook.close()

    def __enter__(self) -> Dataset:
        return self

    def __exit__(self, *_: object) -> None:
        self.close()

    @property
    def headers(self) -> tuple[Any, ...]:
        return tuple(cell.value for cell in self.sheet[1])

    def require_schema(self) -> None:
        if self.headers != HEADERS:
            raise DatasetError(
                "workbook headers do not match the PyResBugs schema; "
                "run 'pyresbugs validate' for details"
            )

    def records(self) -> Iterator[Record]:
        headers = self.headers
        for excel_row, values in enumerate(
            self.sheet.iter_rows(min_row=2, values_only=True), start=2
        ):
            if not any(value is not None for value in values):
                continue
            yield Record(dict(zip(headers, values, strict=True)), excel_row)

    def __len__(self) -> int:
        return sum(1 for _ in self.records())

    def resolve(self, selector: str | int) -> Record:
        text = str(selector).strip()
        records = list(self.records())

        if text.isdecimal():
            ordinal = int(text)
            matches = [record for record in records if record.ordinal == ordinal]
            if matches:
                return matches[0]

        id_match = ID_PATTERN.fullmatch(text)
        if id_match:
            id_fragment = id_match.group(1).upper()
            matches = [record for record in records if record.id[3:].startswith(id_fragment)]
            resolved = self._one_match(text, matches)
            if resolved is not None:
                return resolved

        if SHA_PATTERN.fullmatch(text):
            lowered = text.lower()
            matches = [
                record
                for record in records
                if str(record.values.get("Commit_sha", "")).lower().startswith(lowered)
            ]
            resolved = self._one_match(text, matches)
            if resolved is not None:
                return resolved

        raise RecordNotFoundError(f"no bug matches {text!r}; use 'pyresbugs query' to list bug IDs")

    @staticmethod
    def _one_match(selector: str, matches: list[Record]) -> Record | None:
        if len(matches) == 1:
            return matches[0]
        if len(matches) > 1:
            choices = ", ".join(record.id for record in matches[:8])
            more = " ..." if len(matches) > 8 else ""
            raise AmbiguousSelectorError(
                f"{selector!r} matches {len(matches)} bugs: {choices}{more}"
            )
        return None

    def query(
        self,
        *,
        project: str | None = None,
        fault: str | None = None,
        commit: str | None = None,
    ) -> list[Record]:
        project_lower = project.lower() if project else None
        fault_lower = fault.lower() if fault else None
        commit_lower = commit.lower() if commit else None
        result = []
        for record in self.records():
            values = record.values
            if project_lower and project_lower not in str(values["Project"]).lower():
                continue
            if fault_lower and fault_lower != str(values["Fault_Acronym"]).lower():
                continue
            if commit_lower and not str(values["Commit_sha"]).lower().startswith(commit_lower):
                continue
            result.append(record)
        return result

    def projects(self) -> Counter[str]:
        return Counter(str(record.values["Project"]) for record in self.records())

    def validate(self) -> list[ValidationIssue]:
        issues: list[ValidationIssue] = []
        if self.headers != HEADERS:
            issues.append(
                ValidationIssue(
                    "error",
                    "headers",
                    f"expected {list(HEADERS)!r}, found {list(self.headers)!r}",
                    1,
                )
            )
            return issues

        seen_ids: dict[str, int] = {}
        for record in self.records():
            row_issues = validate_record(record.values, record.excel_row)
            issues.extend(row_issues)
            previous = seen_ids.get(record.id)
            if previous is not None:
                issues.append(
                    ValidationIssue(
                        "error",
                        "duplicate",
                        f"duplicates {record.id} from row {previous}",
                        record.excel_row,
                    )
                )
            else:
                seen_ids[record.id] = record.excel_row
        return issues

    def add(self, values: Mapping[str, Any]) -> Record:
        self.require_schema()
        normalized = normalize_record(values, partial=False)
        issues = validate_record(normalized)
        errors = [issue for issue in issues if issue.level == "error"]
        if errors:
            raise InvalidRecordError("; ".join(str(issue) for issue in errors))

        existing_ids = {record.id for record in self.records()}
        candidate = Record(normalized, self.sheet.max_row + 1)
        if candidate.id in existing_ids:
            raise InvalidRecordError(f"bug {candidate.id} is already in the dataset")

        self.sheet.append([normalized[field] for field in HEADERS])
        return Record(normalized, self.sheet.max_row)

    def update(self, selector: str | int, changes: Mapping[str, Any]) -> tuple[Record, Record]:
        self.require_schema()
        original = self.resolve(selector)
        partial = normalize_record(changes, partial=True)
        merged = {**original.values, **partial}
        issues = validate_record(merged, original.excel_row)
        errors = [issue for issue in issues if issue.level == "error"]
        if errors:
            raise InvalidRecordError("; ".join(str(issue) for issue in errors))

        updated = Record(merged, original.excel_row)
        duplicate = next(
            (
                record
                for record in self.records()
                if record.excel_row != original.excel_row and record.id == updated.id
            ),
            None,
        )
        if duplicate:
            raise InvalidRecordError(
                f"update would duplicate {duplicate.id} at row {duplicate.excel_row}"
            )

        for column, field in enumerate(HEADERS, start=1):
            self.sheet.cell(original.excel_row, column, merged[field])
        return original, updated

    def remove(self, selector: str | int) -> Record:
        self.require_schema()
        record = self.resolve(selector)
        self.sheet.delete_rows(record.excel_row, 1)
        return record

    def save(self, *, backup: str | Path | None = None) -> None:
        errors = [issue for issue in self.validate() if issue.level == "error"]
        if errors:
            preview = "; ".join(str(issue) for issue in errors[:5])
            more = f"; and {len(errors) - 5} more" if len(errors) > 5 else ""
            raise DatasetError(f"refusing to save an invalid dataset: {preview}{more}")

        if backup is not None:
            backup_path = Path(backup)
            if backup_path.resolve() == self.path.resolve():
                raise DatasetError("backup path must differ from the dataset path")
            backup_path.parent.mkdir(parents=True, exist_ok=True)
            shutil.copy2(self.path, backup_path)

        self.path.parent.mkdir(parents=True, exist_ok=True)
        descriptor, temporary_name = tempfile.mkstemp(
            prefix=f".{self.path.stem}-", suffix=self.path.suffix, dir=self.path.parent
        )
        os.close(descriptor)
        temporary_path = Path(temporary_name)
        try:
            self.workbook.save(temporary_path)
            check = load_workbook(temporary_path, read_only=True, data_only=False)
            try:
                if SHEET_NAME not in check.sheetnames:
                    raise DatasetError("temporary workbook failed verification")
                check_headers = tuple(cell.value for cell in check[SHEET_NAME][1])
                if check_headers != HEADERS:
                    raise DatasetError("temporary workbook has invalid headers")
            finally:
                check.close()
            os.replace(temporary_path, self.path)
        finally:
            temporary_path.unlink(missing_ok=True)

    def export_json(self, output: TextIO, records: Iterable[Record] | None = None) -> None:
        selected = list(records if records is not None else self.records())
        json.dump(
            [record.as_dict() for record in selected],
            output,
            ensure_ascii=False,
            indent=2,
        )
        output.write("\n")

    def export_csv(self, output: TextIO, records: Iterable[Record] | None = None) -> None:
        writer = csv.DictWriter(output, fieldnames=("Bug_ID", *HEADERS), lineterminator="\n")
        writer.writeheader()
        for record in records if records is not None else self.records():
            writer.writerow(record.as_dict())


def normalize_record(values: Mapping[str, Any], *, partial: bool) -> dict[str, Any]:
    unknown = sorted(set(values) - set(HEADERS))
    if unknown:
        raise InvalidRecordError(f"unknown field(s): {', '.join(unknown)}")
    if partial:
        if not values:
            raise InvalidRecordError("the update contains no fields")
        return dict(values)
    return {field: values.get(field) for field in HEADERS}


def validate_record(
    values: Mapping[str, Any], excel_row: int | None = None
) -> list[ValidationIssue]:
    issues: list[ValidationIssue] = []
    for field in REQUIRED_FIELDS:
        value = values.get(field)
        if value is None or (isinstance(value, str) and not value.strip()):
            issues.append(ValidationIssue("error", "required", f"{field!r} is empty", excel_row))

    for field, value in values.items():
        if value is not None and not isinstance(value, str):
            issues.append(
                ValidationIssue("error", "type", f"{field!r} must be a string or null", excel_row)
            )

    sha = values.get("Commit_sha")
    if isinstance(sha, str) and sha and not SHA_PATTERN.fullmatch(sha):
        issues.append(
            ValidationIssue(
                "error", "commit", "Commit_sha must contain 7 to 40 hex digits", excel_row
            )
        )

    for field in ("Commit_URL", "Url"):
        value = values.get(field)
        if isinstance(value, str) and value:
            parsed = urlparse(value)
            if parsed.scheme not in ("http", "https") or not parsed.netloc:
                issues.append(
                    ValidationIssue("error", "url", f"{field!r} must be an HTTP(S) URL", excel_row)
                )
    return issues


def load_record_file(path: str | Path) -> dict[str, Any]:
    source = Path(path)
    try:
        data = json.loads(source.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as error:
        raise InvalidRecordError(f"cannot read {source}: {error}") from error
    if not isinstance(data, dict):
        raise InvalidRecordError("a bug record must be a JSON object")
    data.pop("Bug_ID", None)
    return data


def empty_record() -> dict[str, Any]:
    return {field: None if field in OPTIONAL_FIELDS else "" for field in HEADERS}
